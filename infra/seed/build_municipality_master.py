"""総務省「全国地方公共団体コード」Excel から Citify の自治体マスタ CSV を生成する。

Phase 1 (base): 総務省 Excel → 1,794 自治体の骨組み
Phase 2 (supplements): tier1_supplements.csv を municipality_code でマージ
                       → Tier 1 自治体の scraper_type / scraper_base_url 等を上書き

入力:
    --input          総務省 Excel (R6.1.1現在の団体 シート)
    --output         出力 CSV パス
    --supplements    手動補完 CSV (任意、デフォルトは同階層の tier1_supplements.csv)

使用例:
    python infra/seed/build_municipality_master.py \\
        --input /tmp/citify-week0/soumu/000925835.xlsx \\
        --output infra/seed/municipality_master.csv

期待出力行数: 1796 行 (ヘッダ 1 + 国会 1 + 都道府県 47 + 市区町村 1,747)
"""
from __future__ import annotations

import argparse
import csv
import logging
import sys
import unicodedata
from pathlib import Path
from typing import Final

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# 総務省 Excel のメインシート名
SHEET_NAME: Final[str] = "R6.1.1現在の団体"

# 期待されるヘッダ（カラム順検証用、`\n` 改行含む）
EXPECTED_HEADER: Final[tuple[str, ...]] = (
    "団体コード",
    "都道府県名\n（漢字）",
    "市区町村名\n（漢字）",
    "都道府県名\n（カナ）",
    "市区町村名\n（カナ）",
)

# 国会レコード（議事録ソースとしての特殊エントリ）
KOKKAI_RECORD: Final[dict[str, str]] = {
    "municipality_code": "00000",
    "name": "国会",
    "prefecture": "国",
    "kana": "コッカイ",
    "population": "",
    "scraper_type": "kokkai",
    "scraper_base_url": "",
    "tenant_id": "",
    "press_rss_url": "",
    "opendata_url": "",
    "tier": "1",
    "is_active": "true",
    "notes": "国会会議録 (kokkai.ndl.go.jp/api/speech)",
}

# 出力 CSV のカラム順 (Phase 2 で scraper_base_url 追加)
OUTPUT_COLUMNS: Final[tuple[str, ...]] = (
    "municipality_code",
    "name",
    "prefecture",
    "kana",
    "population",
    "scraper_type",
    "scraper_base_url",
    "tenant_id",
    "press_rss_url",
    "opendata_url",
    "tier",
    "is_active",
    "notes",
)

# supplements で上書き可能なカラム（identity 系の name/prefecture/kana/population は対象外）
SUPPLEMENT_OVERRIDE_FIELDS: Final[tuple[str, ...]] = (
    "scraper_type",
    "scraper_base_url",
    "tenant_id",
    "press_rss_url",
    "opendata_url",
    "tier",
    "is_active",
)


def normalize_kana(value: str | None) -> str:
    """半角カナ → 全角カナの正規化（NFKC）。"""
    if not value:
        return ""
    return unicodedata.normalize("NFKC", value)


def truncate_code(code_full: str | int | None) -> str:
    """総務省 6 桁コード（チェックデジット込）の頭 5 桁を返す。"""
    if code_full is None or code_full == "":
        return ""
    s = str(code_full).strip()
    if s.isdigit() and len(s) < 6:
        s = s.zfill(6)
    if len(s) != 6:
        logger.warning("自治体コード長が想定外: %r (len=%d)", s, len(s))
        return s
    return s[:5]


def parse_soumu_xlsx(input_path: Path) -> list[dict[str, str]]:
    """総務省 xlsx のメインシートを Citify スキーマに変換する。"""
    logger.info("入力ファイル読み込み: %s", input_path)
    wb = load_workbook(input_path, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"想定シート '{SHEET_NAME}' が見つかりません。検出シート: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)

    # ヘッダ検証
    header = next(rows_iter)
    header_clean = tuple((c or "").strip() if isinstance(c, str) else (c or "") for c in header[: len(EXPECTED_HEADER)])
    if header_clean != EXPECTED_HEADER:
        raise RuntimeError(
            f"ヘッダ不一致。\n  期待: {EXPECTED_HEADER}\n  実際: {header_clean}"
        )

    records: list[dict[str, str]] = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue

        code_full = raw[0]
        pref_kanji = (raw[1] or "").strip() if isinstance(raw[1], str) else ""
        muni_kanji = (raw[2] or "").strip() if isinstance(raw[2], str) else ""
        pref_kana = (raw[3] or "").strip() if isinstance(raw[3], str) else ""
        muni_kana = (raw[4] or "").strip() if isinstance(raw[4], str) else ""

        if not code_full or not pref_kanji:
            logger.warning("必須欠落のためスキップ: %r", raw[:5])
            continue

        is_prefecture_only = not muni_kanji

        records.append({
            "municipality_code": truncate_code(code_full),
            "name": pref_kanji if is_prefecture_only else muni_kanji,
            "prefecture": pref_kanji,
            "kana": normalize_kana(pref_kana if is_prefecture_only else muni_kana),
            "population": "",
            "scraper_type": "unknown",
            "scraper_base_url": "",
            "tenant_id": "",
            "press_rss_url": "",
            "opendata_url": "",
            "tier": "3",
            "is_active": "false",
            "notes": "prefecture_aggregate" if is_prefecture_only else "",
        })

    logger.info("Excel から %d 件の自治体レコードを抽出", len(records))
    return records


def load_supplements(path: Path) -> dict[str, dict[str, str]]:
    """tier1_supplements.csv を読んで municipality_code をキーとした dict にして返す。"""
    if not path.exists():
        logger.warning("supplements ファイル未配置のためスキップ: %s", path)
        return {}

    supplements: dict[str, dict[str, str]] = {}
    with path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = (row.get("municipality_code") or "").strip()
            if not code:
                continue
            supplements[code] = {k: (v or "").strip() for k, v in row.items()}

    logger.info("supplements を %d 件読み込み: %s", len(supplements), path)
    return supplements


def apply_supplements(
    records: list[dict[str, str]],
    supplements: dict[str, dict[str, str]],
) -> tuple[int, list[str]]:
    """records に supplements を municipality_code 一致で上書きマージ。

    Returns:
        (applied_count, unmatched_codes): 適用件数と、base にマッチしなかった code 一覧
    """
    base_codes = {r["municipality_code"] for r in records}
    applied = 0
    for record in records:
        code = record["municipality_code"]
        if code not in supplements:
            continue
        supp = supplements[code]
        for field in SUPPLEMENT_OVERRIDE_FIELDS:
            value = supp.get(field, "")
            if value:
                record[field] = value
        # notes は base と supplements を結合（両方非空のときのみ "; " 区切り）
        supp_notes = supp.get("notes", "")
        if supp_notes:
            base_notes = record.get("notes", "")
            record["notes"] = f"{base_notes}; {supp_notes}" if base_notes else supp_notes
        applied += 1

    unmatched = sorted(set(supplements.keys()) - base_codes)
    if unmatched:
        logger.warning(
            "supplements の以下 %d 件は base にマッチしなかった: %s",
            len(unmatched), unmatched,
        )

    return applied, unmatched


def write_csv(records: list[dict[str, str]], output_path: Path) -> None:
    """国会レコードを先頭に挿入して CSV を出力する。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    all_records = [KOKKAI_RECORD] + records

    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(OUTPUT_COLUMNS))
        writer.writeheader()
        writer.writerows(all_records)

    logger.info("出力: %s (%d 行 + ヘッダ)", output_path, len(all_records))


def validate(
    records: list[dict[str, str]],
    supplements_applied: int,
    unmatched_codes: list[str],
) -> None:
    """検収用サマリを stdout に表示する。"""
    n_pref = sum(1 for r in records if "prefecture_aggregate" in (r.get("notes") or ""))
    n_muni = len(records) - n_pref

    # scraper_type ごとの集計
    type_counts: dict[str, int] = {}
    for r in records:
        st = r["scraper_type"]
        type_counts[st] = type_counts.get(st, 0) + 1

    print("=" * 60)
    print("検収サマリ")
    print("=" * 60)
    print(f"都道府県全体行 (prefecture_aggregate): {n_pref}")
    print(f"市区町村行: {n_muni}")
    print(f"自治体合計: {len(records)}  (国会を加えると {len(records) + 1})")

    print(f"\nsupplements 適用: {supplements_applied} 件")
    if unmatched_codes:
        print(f"⚠️ supplements マッチ不能: {len(unmatched_codes)} 件 → {unmatched_codes[:5]}...")
    else:
        print("✅ supplements は全件 base にマッチ")

    print("\nscraper_type 別カウント:")
    for st in sorted(type_counts.keys()):
        print(f"  {st:<25} {type_counts[st]:>4}")

    bad_codes = [r["municipality_code"] for r in records if len(r["municipality_code"]) != 5]
    if bad_codes:
        print(f"\n⚠️ コード長が 5 桁でないレコード: {len(bad_codes)} 件、サンプル: {bad_codes[:5]}")
    else:
        print("\n✅ 全レコードのコードが 5 桁")

    # Tier 1 サンプル表示
    tier1 = [r for r in records if r["tier"] == "1"]
    print(f"\nTier 1 自治体: {len(tier1)} 件 (国会含めると {len(tier1) + 1})")
    print("Tier 1 サンプル 5 件:")
    for r in tier1[:5]:
        print(f"  {r['municipality_code']} | {r['name']:<10} | {r['scraper_type']:<22} | {r['scraper_base_url'][:50]}")


def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    parser = argparse.ArgumentParser(
        description="総務省 xlsx + tier1_supplements.csv から municipality_master.csv を生成",
    )
    parser.add_argument("--input", required=True, type=Path, help="総務省 Excel ファイルパス")
    parser.add_argument("--output", required=True, type=Path, help="出力 CSV ファイルパス")
    parser.add_argument(
        "--supplements",
        type=Path,
        default=Path(__file__).parent / "tier1_supplements.csv",
        help="手動補完 CSV (任意、デフォルトは infra/seed/tier1_supplements.csv)",
    )
    args = parser.parse_args()

    if not args.input.exists():
        print(f"ERROR: 入力ファイルが存在しません: {args.input}", file=sys.stderr)
        return 1

    # Phase 1: 総務省 xlsx から base レコード抽出
    records = parse_soumu_xlsx(args.input)

    # Phase 2: tier1_supplements.csv をマージ
    supplements = load_supplements(args.supplements)
    applied, unmatched = apply_supplements(records, supplements)

    # 検収
    validate(records, applied, unmatched)

    # 出力
    write_csv(records, args.output)
    print(f"\n✅ 完了: {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
